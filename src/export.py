"""
export.py

Handles exporting classified transactions into a formatted Excel workbook
based on the schema defined in spreadsheet_schema.py.
"""
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.spreadsheet_schema import get_schema

class SpreadsheetExporter:
    """
    Builds and populates the bookkeeping spreadsheet.
    """
    
    wb: Workbook
    ws: Worksheet

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active # type: ignore
        self.ws.title = "Bookkeeping"
        self.highlight_fill = PatternFill("solid", fgColor="FFFF99")
        self.ignore_fill = PatternFill("solid", fgColor="D9D9D9")
        self.thin_border = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000")
        )
        # Font style for credit-card-sourced transactions
        self.credit_card_font = Font(color="FF0000")  # red


    def set_tax_year(self, tax_year: int | str):
        """
        Write the tax year into A1 with prominent styling (call before build_headers).
        
        args:
            tax_year: int or str tax year (e.g., 2025)
        """
        year_text = str(tax_year)
        cell = self.ws["A1"]
        cell.value = year_text
        cell.font = Font(bold=True, size=18, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="305496")
        # Improve visibility
        self.ws.row_dimensions[1].height = 23
        if (self.ws.column_dimensions["A"].width or 0) < 15:
            self.ws.column_dimensions["A"].width = 15
            

    def build_headers(self):
        """Create header row using schema definitions."""
        schema = get_schema()
        for idx, col in enumerate(schema, start=1):
            self.ws.column_dimensions[col.letter].width = col.width if col.width else 15.0
            cell = self.ws.cell(row=3, column=idx, value=col.name)
            # Apply header style (simplified for now)
            cell.font = Font(bold=True, name="Calibri", size=12, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="9BC2E6")
            thin = Side(border_style="thin", color="000000")
            thick = Side(border_style="thick", color="000000")
            double = Side(border_style="double", color="000000")
            cell.border = Border(top=double, bottom=thick, left=thin, right=thin)


    def write_transaction(self, row_idx: int, transaction: dict):
        """
        Write a single classified transaction into the spreadsheet.
        transaction: dict with keys matching schema column names.
        
        args:
            row_idx: int row number to write to (1-based)
            transaction: dict mapping schema column names -> values
        """
        schema = get_schema()
        is_credit_card_source = bool(transaction.get("credit_card_source"))
        
        for idx, col in enumerate(schema, start=1):
            value = transaction.get(col.name)
            cell = self.ws.cell(row=row_idx, column=idx)
            if col.formula_template and value is None:
                # Insert formula dynamically
                formula = col.formula_template.format(row=row_idx)
                cell.value = formula # type: ignore
            else:
                cell.value = value
                
            # Apply number format based on data_format
            if col.data_format["number"]["category"] == "Currency":
                cell.number_format = '$#,##0.00'
            elif col.data_format["number"]["category"] == "Date":
                cell.number_format = 'YYYY-MM-DD'
                
            # Apply highlighting and borders if manual review flag is set
            if transaction.get("highlight"):
                cell.fill = self.highlight_fill
            if transaction.get("ignore"):
                cell.fill = self.ignore_fill
            
            # Credit-card source visual cue (red font)
            if is_credit_card_source:
                # Merge with any existing style by copying current font and changing color
                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=self.credit_card_font.color,
                )
                    
            cell.border = self.thin_border


    def finalize_totals_row(self, start_row: int, end_row: int):
        """
        Add totals row at the bottom for numeric columns (C through AA).
        
        args:
            start_row: int first data row index
            end_row: int last data row index
        """
        schema = get_schema()
        totals_row = end_row + 1
        
        for idx, col in enumerate(schema, start=1):
            if (col.letter >= "C" and col.letter <= "Z") or (col.letter == "AA"):
                formula = f"=SUM({col.letter}{start_row}:{col.letter}{end_row})"
                cell = self.ws.cell(row=totals_row, column=idx)
                cell.value = formula # type: ignore
                # Apply number format
                if col.data_format["number"]["category"] == "Currency":
                    cell.number_format = '$#,##0.00'
                # Apply bold font for visibility
                cell.font = Font(bold=True)
            elif col.name == "Item":
                # Add "Totals" label for visibility
                cell = self.ws.cell(row=totals_row, column=idx)
                cell.value = "Totals" # type: ignore
                cell.font = Font(bold=True)
                
        # Add borders to all cells in the totals row for visibility
        for idx in range(1, len(schema) + 1):
            cell = self.ws.cell(row=totals_row, column=idx)
            cell.border = self.thin_border


    def add_color_legend(self, last_transaction_row: int, separation: int = 5):
        """
        Append a two-column color guide starting `separation` empty rows after the
        last_transaction_row. Left column (A) is a single merged cell (4 rows tall)
        with centered 'COLOR GUIDE', right column (B) contains 4 descriptive rows.
        
        args:
            last_transaction_row: int last row index containing transaction data
            separation: int number of empty rows between last transaction and legend (default: 5)
        """
        start = last_transaction_row + separation + 1
        end = start + 3  # 4 rows tall

        # Merge left column A over the 4 rows
        left_range = f"A{start}:A{end}"
        self.ws.merge_cells(left_range)
        left_cell = self.ws[f"A{start}"]
        left_cell.value = "COLOR GUIDE"
        left_cell.font = Font(bold=True, size=11)
        left_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Prepare right column texts + styles
        rows = [
            ("BLACK - BANK ACCOUNT - DEBIT (CHEQUING)", {"font_color": "000000", "fill": None}),
            ("GREEN - PAID CASH", {"font_color": "00A000", "fill": None}),
            ("RED - PAID OR RELATED TO CREDIT CARDS", {"font_color": "FF0000", "fill": None}),
            ("HIGHLIGHTED YELLOW - Manual review required", {"font_color": "000000", "fill": "FFFF00"}),
        ]

        for idx, (text, style) in enumerate(rows):
            r = start + idx
            cell = self.ws[f"B{r}"]
            cell.value = text
            cell.font = Font(bold=True, color=style["font_color"])
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if style["fill"]:
                cell.fill = PatternFill("solid", fgColor=style["fill"])
            self.ws.row_dimensions[r].height = 18


    def save(self, filename: str):
        """Save the workbook to disk."""
        self.wb.save(filename)

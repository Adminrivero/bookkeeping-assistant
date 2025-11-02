import pytest
from openpyxl import Workbook
from src.pipeline import run_pipeline

@pytest.fixture
def fake_rules(tmp_path):
    # Minimal ruleset for testing
    rules = {
        "_rules": [
            {
                "category_name": "Test Coffee",
                "transaction_type": "EXPENSE",
                "logic": "MUST_MATCH_ANY",
                "rules": [
                    {"field": "Description", "operator": "CONTAINS", "value": "COFFEE"}
                ],
                "dual_entry": {
                    "DR_COLUMN": {"name": "Food Expenses from Business Meetings", "letter": "T"},
                    "CR_COLUMN": {"name": "Shareholder Contribution (CR)", "letter": "F"},
                    "APPLY_PERCENTAGE": 1.0
                }
            }
        ]
    }
    path = tmp_path / "rules.json"
    import json
    with open(path, "w") as f:
        json.dump(rules, f)
    return str(path)

@pytest.fixture
def fake_transactions():
    return [
        {"Date": "2025-01-01", "Description": "Morning COFFEE", "Debit": 5.00, "Credit": ""},
        {"Date": "2025-01-02", "Description": "Unknown Vendor", "Debit": 20.00, "Credit": ""}
    ]

def test_run_pipeline_creates_workbook(fake_transactions, fake_rules):
    wb = run_pipeline(fake_transactions, fake_rules)
    assert isinstance(wb, Workbook)

    ws = wb.active
    # Check headers
    assert ws["A3"].value == "Date" # type: ignore
    assert ws["B3"].value == "Item" # type: ignore

    # Check first transaction classified as "Test Coffee"
    assert ws["A4"].value == "2025-01-01" # type: ignore
    assert ws["B4"].value == "Morning COFFEE" # type: ignore
    # Debit should appear in DR column "T"
    assert ws["T4"].value == 5.00 # type: ignore
    # Credit should appear in CR column "F"
    assert ws["F4"].value == -5.00 # type: ignore

    # Second transaction unclassified → should land in Notes
    assert "unclassified" in ws["AB5"].value # type: ignore
